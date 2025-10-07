import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from decimal import Decimal
from Pedidos.models import Pedido
from Productos.models import Producto
from django.shortcuts import get_object_or_404

def exportar_pedidos_excel(request):
    # Crear el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos y Productos"

    # Encabezado
    columnas = [
        "Código Pedido", "Cliente", "Fecha Pedido", "Total Neto", "IVA", "Total Pedido",
        "Código Orden Compra", "Producto", "Tipo", "Tamaño", "Observación", 
        "Cantidad", "Precio Unitario", "Subtotal"
    ]
    ws.append(columnas)

    # Estilo de encabezado
    for col in range(1, len(columnas) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="3767a6", end_color="3767a6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Agregar datos
    fila = 2
    pedidos = Pedido.objects.prefetch_related('producto_set', 'cliente').all().order_by('-fecha_pedido')

    for pedido in pedidos:
        productos = pedido.producto_set.all()
        if productos.exists():
            for producto in productos:
                ws.append([
                    pedido.cod_pedido,
                    pedido.cliente.razon_social_cliente if pedido.cliente else "—",
                    pedido.fecha_pedido.strftime("%d-%m-%Y") if pedido.fecha_pedido else "",
                    float(pedido.total_neto_pedido),
                    float(pedido.iva_pedido),
                    float(pedido.total_pedido),
                    pedido.cod_orden_compra or "",
                    producto.nombre_producto,
                    producto.tipo_producto or "",
                    producto.tamano_producto or "",
                    producto.observacion_producto or "",
                    producto.cantidad_producto,
                    float(producto.precio_unitario_producto),
                    float(producto.subtotal()),
                ])
                fila += 1
        else:
            # Pedido sin productos
            ws.append([
                pedido.cod_pedido,
                pedido.cliente.razon_social_cliente if pedido.cliente else "—",
                pedido.fecha_pedido.strftime("%d-%m-%Y") if pedido.fecha_pedido else "",
                float(pedido.total_neto_pedido),
                float(pedido.iva_pedido),
                float(pedido.total_pedido),
                pedido.cod_orden_compra or "",
                "—", "", "", "", "", "", "",
            ])
            fila += 1

    # Ajustar ancho de columnas automáticamente
    for columna in ws.columns:
        max_length = 0
        columna_letra = columna[0].column_letter
        for cell in columna:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[columna_letra].width = max_length + 2

    # Retornar archivo Excel como descarga
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="pedidos_productos.xlsx"'
    wb.save(response)
    return response



def exportar_pedido_excel(request, pedido_id):
    # Obtener el pedido o lanzar 404 si no existe
    pedido = get_object_or_404(Pedido, id=pedido_id)
    productos = pedido.producto_set.all()

    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Pedido {pedido.cod_pedido}"

    # Escribir encabezados de pedido
    ws.append(["Código Pedido", "Código Orden Compra", "Cliente", "Fecha Pedido", "Total Neto", "IVA", "Total"])
    ws.append([
        pedido.cod_pedido,
        pedido.cod_orden_compra,
        pedido.cliente.razon_social_cliente,
        pedido.fecha_pedido.strftime("%Y-%m-%d"),
        float(pedido.total_neto_pedido),
        float(pedido.iva_pedido),
        float(pedido.total_pedido)
    ])

    # Espacio en blanco
    ws.append([])

    # Encabezados de productos
    ws.append(["Nombre Producto", "Tipo", "Tamaño", "Observación", "Cantidad", "Precio Unitario", "Subtotal"])

    # Escribir cada producto
    for p in productos:
        ws.append([
            p.nombre_producto,
            p.tipo_producto,
            p.tamano_producto,
            p.observacion_producto,
            p.cantidad_producto,
            float(p.precio_unitario_producto),
            float(p.subtotal())
        ])

    # Preparar respuesta HTTP con Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Pedido_{pedido.cod_pedido}.xlsx'
    wb.save(response)
    return response